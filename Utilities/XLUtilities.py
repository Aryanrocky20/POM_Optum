import openpyxl


def getRowCount(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_row


def getColumnCount(fileName, sheetName):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.max_column


def readData(fileName, sheetName, row, column):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=row, column=column)


def writeData(fileName, sheetName, row, column, data):
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.get_sheet_by_name(sheetName)
    sheet.cell(row=row, column=column).value = data
    workbook.save(fileName)
